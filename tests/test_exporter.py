import pytest
from pathlib import Path
from openpyxl import load_workbook
from src.exporter import tree_to_rows, export_trees_to_xlsx
from src.models import RequirementNode, RequirementTree

def _make_tree():
    child = RequirementNode(
        level=2, level_name="Major System", allocation="SDS",
        chapter_code="SDS-Ch 2", derived_name="Propulsion",
        technical_requirement="The Propulsion System shall provide power.",
        rationale="R.", system_hierarchy_id="SBS 200",
        verification_method=["Analysis"], verification_event=["Design Review"],
        test_case_descriptions=["Review docs."], acceptance_criteria="Phase 1 review.")
    root = RequirementNode(
        level=1, level_name="Whole Ship", allocation="GTR",
        chapter_code="GTR-Ch 3", derived_name="Speed",
        technical_requirement="The Vessel shall achieve 17 knots.",
        rationale="R.", system_hierarchy_id="SBS 700",
        verification_method=["Test"], verification_event=["Sea Trials"],
        test_case_descriptions=["Speed trial."], acceptance_criteria="Phase 5 sea trial.",
        children=[child])
    return RequirementTree(dig_id="9584", dig_text="Ship speed.", root=root)

class TestTreeToRows:
    def test_flattens_tree(self):
        rows = tree_to_rows(_make_tree())
        assert len(rows) == 2

    def test_row_has_all_columns(self):
        rows = tree_to_rows(_make_tree())
        row = rows[0]
        for key in ["dig_id", "level", "node_id", "parent_id", "technical_requirement", "verification_method"]:
            assert key in row

    def test_parent_child_relationship(self):
        rows = tree_to_rows(_make_tree())
        assert rows[1]["parent_id"] == rows[0]["node_id"]

    def test_verification_arrays_joined(self):
        rows = tree_to_rows(_make_tree())
        assert isinstance(rows[0]["verification_method"], str)

class TestExportXlsx:
    def test_creates_xlsx_file(self, tmp_path):
        export_trees_to_xlsx([_make_tree()], tmp_path / "test.xlsx")
        assert (tmp_path / "test.xlsx").exists()

    def test_xlsx_has_correct_rows(self, tmp_path):
        export_trees_to_xlsx([_make_tree()], tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        assert wb.active.max_row == 3
        wb.close()
